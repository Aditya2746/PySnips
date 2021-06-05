""" Imports """
import openpyxl as xl
from openpyxl.chart import BarChart,PieChart,Reference 

workbook = xl.load_workbook("excel\Data.xlsx")# loading workbook
sheet = workbook["Sheet1"] # accessing sheet 1
max_row = sheet.max_row + 1 # getting maximum number of rows + 1

for row in range(2, max_row):
    cell = sheet.cell(row, 4)
    corrected_marks = cell.value + 5
    corrected_marks_cell = sheet.cell(row, 4)
    corrected_marks_cell.value = corrected_marks

values = Reference(sheet, min_row=2, max_row=sheet.max_row, max_col=6, min_col=4)

chart = PieChart()
chart.add_data(values) 
sheet.add_chart(chart, 'e2')

workbook.save('excel\Data.xlsx')
